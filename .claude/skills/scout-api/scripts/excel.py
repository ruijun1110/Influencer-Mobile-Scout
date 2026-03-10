#!/usr/bin/env python3
"""
excel.py — xlsx read/write helpers for influencer scouting.
All functions operate on data/influencers.xlsx relative to project_root.
"""
import re
import statistics
from datetime import date
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    import os; os.system('pip3 install openpyxl -q')
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

# ---------------------------------------------------------------------------
# Sheet schemas — order matters for new sheets; existing cols are preserved
# ---------------------------------------------------------------------------
CANDIDATES_COLS = [
    'handle', 'triggering_video_url', 'triggering_play_count',
    'keyword', 'campaign', 'audit_status'
]
INFLUENCERS_COLS = [
    'handle', 'profile_url', 'max_views', 'min_views', 'median_views',
    'triggering_video_url', 'triggering_play_count',
    'keyword', 'campaign', 'scouted_date', 'notes'
]
SEARCH_LOG_COLS = [
    'keyword', 'results_checked', 'candidates_found', 'qualified',
    'duration_mins', 'campaign', 'run_date'
]
CONFIG_COLS = ['key', 'value']

PROJECT_ROOT = Path(__file__).resolve().parents[4]
XLSX_PATH = PROJECT_ROOT / 'data' / 'influencers.xlsx'


def _style_header(cell):
    cell.font = Font(bold=True)
    cell.fill = PatternFill('solid', start_color='D9E1F2')
    cell.alignment = Alignment(horizontal='center')


def _init_xlsx():
    """Create influencers.xlsx with correct scout-api schemas from scratch."""
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    def make(name, cols, widths):
        ws = wb.create_sheet(name)
        for i, h in enumerate(cols, 1):
            c = ws.cell(row=1, column=i, value=h)
            _style_header(c)
        for i, w in enumerate(widths):
            ws.column_dimensions[ws.cell(row=1, column=i + 1).column_letter].width = w
        return ws

    make('Influencers', INFLUENCERS_COLS,
         [22, 45, 14, 14, 14, 50, 22, 30, 20, 14, 30])
    make('Candidates', CANDIDATES_COLS,
         [22, 50, 16, 30, 20, 16])
    make('Search Log', SEARCH_LOG_COLS,
         [30, 18, 18, 12, 16, 20, 14])

    ws_cfg = make('Config', CONFIG_COLS, [28, 20])
    for row, (k, v) in enumerate([
        ('view_threshold', 10000),
        ('min_video_views', 5000),
        ('avg_view_standard', 50000),
        ('recent_video_count', 10),
        ('max_candidates_per_keyword', 5),
    ], 2):
        ws_cfg.cell(row=row, column=1, value=k)
        ws_cfg.cell(row=row, column=2, value=v)

    wb.save(XLSX_PATH)
    print(f'Created {XLSX_PATH}')


def _get_wb_ws(sheet_name: str, required_cols: list):
    """Open xlsx and return (wb, ws). Creates xlsx/sheet with headers if missing."""
    if not XLSX_PATH.exists():
        _init_xlsx()
    wb = load_workbook(XLSX_PATH)

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        for i, col in enumerate(required_cols, 1):
            c = ws.cell(row=1, column=i, value=col)
            _style_header(c)
    else:
        ws = wb[sheet_name]
        _ensure_columns(ws, required_cols)

    return wb, ws


def _ensure_columns(ws, required_cols: list):
    """Append missing columns to header row — never drops existing."""
    if ws.max_row == 0:
        return
    existing = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for col_name in required_cols:
        if col_name not in existing:
            next_col = ws.max_column + 1
            c = ws.cell(row=1, column=next_col, value=col_name)
            _style_header(c)
            existing.append(col_name)


def _col_index(ws, name: str):
    """Return 1-based column index by header name, or None."""
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == name:
            return c
    return None


def _row_to_dict(ws, row_num: int) -> dict:
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    return {h: ws.cell(row=row_num, column=i + 1).value for i, h in enumerate(headers) if h}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_config(campaign_name: str) -> dict:
    """Read Config sheet defaults, overlay campaign.md YAML front-matter."""
    config = {}

    # --- Config sheet ---
    if XLSX_PATH.exists():
        wb = load_workbook(XLSX_PATH, read_only=True)
        if 'Config' in wb.sheetnames:
            ws = wb['Config']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    try:
                        config[row[0]] = int(row[1]) if str(row[1]).isdigit() else row[1]
                    except (TypeError, ValueError):
                        config[row[0]] = row[1]
        wb.close()

    # --- campaign.md YAML front-matter ---
    campaign_path = PROJECT_ROOT / 'context' / 'campaigns' / campaign_name / 'campaign.md'
    if campaign_path.exists():
        text = campaign_path.read_text()
        if text.startswith('---'):
            fm_block = text.split('---')[1]
            for line in fm_block.splitlines():
                line = line.strip()
                if ':' in line and not line.startswith('#'):
                    key, _, val = line.partition(':')
                    val = val.strip().strip('"').strip("'")
                    if val.lstrip('-').isdigit():
                        config[key.strip()] = int(val)
                    elif val.replace('.', '', 1).isdigit():
                        config[key.strip()] = float(val)
                    elif val:
                        config[key.strip()] = val

    # Defaults
    config.setdefault('view_threshold', 10000)
    config.setdefault('avg_view_standard', 50000)
    config.setdefault('recent_video_count', 10)
    config.setdefault('max_candidates_per_keyword', 5)
    config.setdefault('min_video_views', config['view_threshold'])
    return config


def append_candidates(rows: list[dict]):
    """Append rows to Candidates sheet; skip if handle+campaign already exists."""
    wb, ws = _get_wb_ws('Candidates', CANDIDATES_COLS)
    existing = set()
    for r in range(2, ws.max_row + 1):
        h = ws.cell(row=r, column=_col_index(ws, 'handle') or 1).value
        camp = ws.cell(row=r, column=_col_index(ws, 'campaign') or 5).value
        if h and camp:
            existing.add((h, camp))

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    added = 0
    for row in rows:
        key = (row.get('handle'), row.get('campaign'))
        if key in existing:
            continue
        next_row = ws.max_row + 1
        for i, h in enumerate(headers, 1):
            if h and h in row:
                ws.cell(row=next_row, column=i, value=row[h])
        existing.add(key)
        added += 1

    wb.save(XLSX_PATH)
    return added


def candidate_exists(handle: str, campaign: str) -> bool:
    if not XLSX_PATH.exists():
        return False
    wb = load_workbook(XLSX_PATH, read_only=True)
    if 'Candidates' not in wb.sheetnames:
        wb.close()
        return False
    ws = wb['Candidates']
    handle_col = _col_index(ws, 'handle') or 1
    camp_col = _col_index(ws, 'campaign') or 5
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=handle_col).value == handle and \
           ws.cell(row=r, column=camp_col).value == campaign:
            wb.close()
            return True
    wb.close()
    return False


def update_candidate_status(handle: str, campaign: str, status: str, notes: str = ''):
    wb, ws = _get_wb_ws('Candidates', CANDIDATES_COLS)
    handle_col = _col_index(ws, 'handle')
    camp_col = _col_index(ws, 'campaign')
    status_col = _col_index(ws, 'audit_status')

    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=handle_col).value == handle and \
           ws.cell(row=r, column=camp_col).value == campaign:
            ws.cell(row=r, column=status_col, value=status)
            notes_col = _col_index(ws, 'notes')
            if notes_col and notes:
                ws.cell(row=r, column=notes_col, value=notes)
            break

    wb.save(XLSX_PATH)


def get_pending_candidates(campaign: str) -> list[dict]:
    if not XLSX_PATH.exists():
        return []
    wb = load_workbook(XLSX_PATH, read_only=True)
    if 'Candidates' not in wb.sheetnames:
        wb.close()
        return []
    ws = wb['Candidates']
    camp_col = _col_index(ws, 'campaign')
    status_col = _col_index(ws, 'audit_status')
    results = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=camp_col).value == campaign and \
           ws.cell(row=r, column=status_col).value in (None, '', 'pending'):
            results.append(_row_to_dict(ws, r))
    wb.close()
    return results


def append_influencer(row: dict):
    """Append to Influencers sheet; skip if handle+campaign already exists."""
    wb, ws = _get_wb_ws('Influencers', INFLUENCERS_COLS)
    handle_col = _col_index(ws, 'handle') or 1
    camp_col = _col_index(ws, 'campaign') or 10
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=handle_col).value == row.get('handle') and \
           ws.cell(row=r, column=camp_col).value == row.get('campaign'):
            wb.close()
            return  # already exists
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    next_row = ws.max_row + 1
    for i, h in enumerate(headers, 1):
        if h and h in row:
            ws.cell(row=next_row, column=i, value=row[h])
    wb.save(XLSX_PATH)


def append_search_log(row: dict):
    wb, ws = _get_wb_ws('Search Log', SEARCH_LOG_COLS)
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    next_row = ws.max_row + 1
    for i, h in enumerate(headers, 1):
        if h and h in row:
            ws.cell(row=next_row, column=i, value=row[h])
    wb.save(XLSX_PATH)


def read_keywords(campaign_name: str) -> list[dict]:
    """Read keywords.md table, return list of {keyword, status, source, date} dicts."""
    kw_path = PROJECT_ROOT / 'context' / 'campaigns' / campaign_name / 'keywords.md'
    if not kw_path.exists():
        return []
    lines = kw_path.read_text().splitlines()
    keywords = []
    in_table = False
    headers = []
    for line in lines:
        line = line.strip()
        if not line.startswith('|'):
            continue
        cols = [c.strip() for c in line.strip('|').split('|')]
        if not headers:
            headers = cols
            in_table = True
            continue
        if set(line.replace('|', '').replace('-', '').replace(' ', '')) == set():
            continue  # separator row
        row = dict(zip(headers, cols))
        keywords.append(row)
    return keywords


def append_keyword(campaign_name: str, keyword: str, source: str = 'imessage') -> bool:
    """Append a new pending keyword row to keywords.md inside the existing table.
    Returns False if keyword already exists (any status), True if appended."""
    kw_path = PROJECT_ROOT / 'context' / 'campaigns' / campaign_name / 'keywords.md'
    if not kw_path.exists():
        return False

    existing = read_keywords(campaign_name)
    for row in existing:
        if row.get('keyword', '').strip().lower() == keyword.strip().lower():
            return False  # duplicate

    today = date.today().isoformat()
    lines = kw_path.read_text().splitlines()

    # Find the last table row (any line starting with |)
    last_table_line = None
    for i, line in enumerate(lines):
        if line.strip().startswith('|'):
            last_table_line = i

    if last_table_line is None:
        # No table found — just append
        kw_path.write_text('\n'.join(lines) + f'\n| {keyword} | pending | {source} | {today} |\n')
        return True

    # Detect column widths from separator row (e.g. | --- | --- |)
    sep_line = next((l for l in lines if l.strip().startswith('|') and set(l.replace('|', '').replace('-', '').replace(' ', '')) == set()), None)
    values = [keyword, 'pending', source, today]
    if sep_line:
        # Width of each col = number of dashes in that cell
        w = [len(c.strip()) for c in sep_line.strip('|').split('|')]
        w = [max(w[i] if i < len(w) else 0, len(v)) for i, v in enumerate(values)]
        new_row = '| ' + ' | '.join(v.ljust(w[i]) for i, v in enumerate(values)) + ' |'
    else:
        new_row = '| ' + ' | '.join(values) + ' |'

    lines.insert(last_table_line + 1, new_row)
    kw_path.write_text('\n'.join(lines) + '\n')
    return True


def mark_keyword_searched(campaign_name: str, keyword: str):
    """Update keyword status to 'searched' in keywords.md."""
    kw_path = PROJECT_ROOT / 'context' / 'campaigns' / campaign_name / 'keywords.md'
    if not kw_path.exists():
        return
    text = kw_path.read_text()
    # Replace the first matching pending row
    lines = text.splitlines()
    for i, line in enumerate(lines):
        if line.strip().startswith('|') and keyword in line and 'pending' in line:
            lines[i] = line.replace('pending', 'searched', 1)
            break
    kw_path.write_text('\n'.join(lines) + '\n')
