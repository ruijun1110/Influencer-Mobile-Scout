#!/usr/bin/env python3
"""Creates data/influencers.xlsx with all required sheets. Safe to re-run."""

import os, sys, argparse
from pathlib import Path

parser = argparse.ArgumentParser()
parser.add_argument('--project-root', default=None)
args = parser.parse_args()

project_root = Path(args.project_root) if args.project_root else Path(__file__).parent.parent.parent.parent.parent
OUTPUT = project_root / 'data' / 'influencers.xlsx'

if OUTPUT.exists():
    print('data/influencers.xlsx already exists — skipping.')
    sys.exit(0)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print('Installing openpyxl...')
    os.system('pip3 install openpyxl -q')
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

def make_sheet(wb, name, headers, widths):
    ws = wb.create_sheet(name)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', start_color='D9E1F2')
        c.alignment = Alignment(horizontal='center')
    for i, w in enumerate(widths):
        ws.column_dimensions[ws.cell(row=1, column=i+1).column_letter].width = w
    return ws

wb = Workbook()
wb.remove(wb.active)

make_sheet(wb, 'Influencers', [
    'handle', 'profile_url', 'max_views', 'min_views', 'median_views',
    'triggering_video_url', 'triggering_play_count',
    'keyword', 'campaign', 'scouted_date', 'notes'
], [22, 45, 14, 14, 14, 50, 22, 30, 20, 14, 30])

make_sheet(wb, 'Candidates', [
    'handle', 'triggering_video_url', 'triggering_play_count', 'keyword', 'campaign', 'audit_status'
], [22, 50, 22, 30, 20, 16])

make_sheet(wb, 'Search Log', [
    'keyword', 'results_checked', 'candidates_found', 'qualified', 'duration_mins', 'campaign', 'run_date'
], [30, 18, 18, 12, 16, 20, 14])

ws_cfg = make_sheet(wb, 'Config', ['key', 'value'], [28, 20])
defaults = [
    ('view_threshold', 10000),
    ('min_video_views', 10000),
    ('avg_view_standard', 50000),
    ('recent_video_count', 10),
    ('max_candidates_per_keyword', 5),
    ('max_concurrent_audits', 10),
]
for row, (k, v) in enumerate(defaults, 2):
    ws_cfg.cell(row=row, column=1, value=k)
    ws_cfg.cell(row=row, column=2, value=v)

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)
print(f'Created {OUTPUT}')
