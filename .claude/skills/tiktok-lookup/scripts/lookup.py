# /// script
# dependencies = ["httpx", "openpyxl", "python-dotenv"]
# ///
"""
lookup.py — TikHub similar user lookup + xlsx write

Usage: uv run lookup.py <handle> <requested_by> <data_dir>
Prints formatted iMessage reply to stdout.
"""

import sys
import os
from pathlib import Path
from datetime import date

# Load .env from .claude/.env (3 levels up from scripts/)
dotenv_path = Path(__file__).resolve().parents[4] / '.claude' / '.env'
if dotenv_path.exists():
    from dotenv import load_dotenv
    load_dotenv(dotenv_path)

import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_URL = "https://api.tikhub.io/api/v1"
MAX_RESULTS = 10


def get_headers():
    api_key = os.environ.get('TIKHUB_API_KEY')
    if not api_key:
        print(
            f"ERROR: TIKHUB_API_KEY not set in {dotenv_path}",
            file=sys.stderr,
        )
        sys.exit(1)
    return {"Authorization": f"Bearer {api_key}"}


def get_sec_uid(handle: str, headers: dict) -> str:
    resp = httpx.get(
        f"{BASE_URL}/tiktok/web/fetch_user_profile",
        params={"uniqueId": handle},
        headers=headers,
        timeout=15,
    )
    resp.raise_for_status()
    payload = resp.json()
    inner = payload.get("data", {})
    status = inner.get("statusCode", -1)
    if status != 0:
        raise ValueError(f"TikTok account @{handle} not found (statusCode={status})")
    sec_uid = inner.get("userInfo", {}).get("user", {}).get("secUid")
    if not sec_uid:
        raise ValueError(f"Could not extract sec_uid for @{handle}")
    return sec_uid


def get_similar_users(sec_uid: str, headers: dict) -> list[dict]:
    resp = httpx.get(
        f"{BASE_URL}/tiktok/app/v3/fetch_similar_user_recommendations",
        params={"sec_uid": sec_uid},
        headers=headers,
        timeout=15,
    )
    resp.raise_for_status()
    data = resp.json()
    user_list = data.get("data", {}).get("users", [])
    return user_list[:MAX_RESULTS]


HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ROW_FILLS = [PatternFill("solid", fgColor="F8F9FA"), PatternFill("solid", fgColor="FFFFFF")]
LINK_FONT = Font(color="0066CC", underline="single")
BORDER_SIDE = Side(style="thin", color="DEE2E6")
CELL_BORDER = Border(bottom=Border(bottom=BORDER_SIDE).bottom)
COL_WIDTHS = [20, 20, 40, 14, 20]  # queried_handle, similar_handle, profile_url, lookup_date, requested_by


def apply_header_style(ws):
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_idx - 1]
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def style_data_row(ws, row_idx: int):
    fill = ROW_FILLS[row_idx % 2]
    for col_idx, cell in enumerate(ws[row_idx], 1):
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
        # Make profile_url column look like a link
        if col_idx == 3:
            cell.font = LINK_FONT


def save_to_xlsx(data_dir: str, queried_handle: str, users: list[dict], requested_by: str):
    xlsx_path = Path(data_dir) / "similar_users.xlsx"
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    is_new = not xlsx_path.exists()
    if is_new:
        wb = Workbook()
        ws = wb.active
        ws.title = "Lookups"
        ws.append(["queried_handle", "similar_handle", "profile_url", "lookup_date", "requested_by"])
        apply_header_style(ws)
    else:
        wb = load_workbook(xlsx_path)
        ws = wb["Lookups"] if "Lookups" in wb.sheetnames else wb.active

    today = str(date.today())
    for u in users:
        similar_handle = u.get("unique_id", "") or u.get("uniqueId", "")
        if similar_handle:
            ws.append([
                queried_handle,
                similar_handle,
                f"https://www.tiktok.com/@{similar_handle}",
                today,
                requested_by,
            ])
            style_data_row(ws, ws.max_row)

    wb.save(xlsx_path)


def format_reply(queried_handle: str, users: list[dict]) -> str:
    handles = [
        u.get("unique_id", "") or u.get("uniqueId", "")
        for u in users
        if u.get("unique_id", "") or u.get("uniqueId", "")
    ]
    if not handles:
        return f"__HEADER__No similar creators found for @{queried_handle}."
    urls = "\n\n".join(f"https://www.tiktok.com/@{h}" for h in handles)
    return f"__HEADER__Similar to @{queried_handle} ({len(handles)} creators):\n__URLS__\n{urls}"


def main():
    if len(sys.argv) < 4:
        print("Usage: lookup.py <handle> <requested_by> <data_dir>", file=sys.stderr)
        sys.exit(1)

    handle = sys.argv[1].lstrip('@')
    requested_by = sys.argv[2]
    data_dir = sys.argv[3]

    headers = get_headers()

    sec_uid = get_sec_uid(handle, headers)
    users = get_similar_users(sec_uid, headers)
    save_to_xlsx(data_dir, handle, users, requested_by)
    print(format_reply(handle, users))


if __name__ == "__main__":
    main()
